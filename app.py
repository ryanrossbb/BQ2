"""
Medical Benefits Comparison — v4 Phase 2
Adds: server-side template library with auto-selection by plan count,
admin upload/swap UI, removed per-session template upload.
"""

import os
import io
import json
import base64
import secrets
import urllib.request
import urllib.error
from functools import wraps
from pathlib import Path

from flask import Flask, request, jsonify, send_file, send_from_directory, session
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

try:
    from pypdf import PdfReader, PdfWriter
    PYPDF_OK = True
except ImportError:
    PYPDF_OK = False

app = Flask(__name__, static_folder="static")
app.secret_key = os.environ.get("FLASK_SECRET", secrets.token_hex(32))

# ─── CONFIG ───────────────────────────────────────────────────────────────────
APP_PASSWORD  = os.environ.get("APP_PASSWORD", "changeme")
ANTHROPIC_KEY = os.environ.get("ANTHROPIC_API_KEY", "")

# Persistent template directory. On Render free tier this resets on redeploys,
# but uploaded templates persist across the running process.
# For permanent storage, mount a Render Disk to this path.
TEMPLATE_DIR = Path(os.environ.get("TEMPLATE_DIR", "/tmp/templates"))
TEMPLATE_DIR.mkdir(parents=True, exist_ok=True)

# Mapping: quote type → filename
TEMPLATE_FILES = {
    "single": TEMPLATE_DIR / "template_single.xlsx",
    "double": TEMPLATE_DIR / "template_double.xlsx",
    "triple": TEMPLATE_DIR / "template_triple.xlsx",
}
PLANS_PER_BUNDLE = {"single": 1, "double": 2, "triple": 3}

def require_auth(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get("authed"):
            return jsonify({"error": {"message": "Not authenticated"}}), 401
        return f(*args, **kwargs)
    return wrapper

def pick_template_path(quote_type):
    """Choose template by quote type: single / double / triple."""
    return TEMPLATE_FILES.get((quote_type or "single").lower())

# ─── PDF UTILITIES ────────────────────────────────────────────────────────────
def filter_pdf_pages(pdf_bytes, page_range_str):
    if not PYPDF_OK or not page_range_str or page_range_str.strip().lower() in ("", "all"):
        return pdf_bytes
    reader = PdfReader(io.BytesIO(pdf_bytes))
    total = len(reader.pages)
    pages = set()
    for part in page_range_str.split(","):
        part = part.strip()
        if "-" in part:
            try:
                s, e = part.split("-", 1)
                pages.update(range(max(1, int(s.strip())), min(total, int(e.strip())) + 1))
            except ValueError:
                pass
        elif part.isdigit():
            p = int(part)
            if 1 <= p <= total:
                pages.add(p)
    if not pages:
        return pdf_bytes
    writer = PdfWriter()
    for p in sorted(pages):
        writer.add_page(reader.pages[p - 1])
    buf = io.BytesIO()
    writer.write(buf)
    return buf.getvalue()

# ─── EXCEL TEMPLATE WRITER ────────────────────────────────────────────────────
LABEL_MAP = {
    "Network Availability":                              "network",
    "Deductible (Employee / Family)":                    "ded",
    "Coinsurance (Member / Carrier)":                    "coinsurance",
    "MOOP (Employee / Family)":                          "moop",
    "PCP Copay":                                         "pcp",
    "Telehealth":                                        "telehealth",
    "Specialist Copay":                                  "specialist",
    "Inpatient Hospitalization":                         "inpatient",
    "Outpatient Facility":                               "op_facility",
    "Emergency Room":                                    "er",
    "Urgent Care":                                       "urgent_care",
    "Laboratory":                                        "lab",
    "X-ray":                                             "xray",
    "Imaging (CT, MRI, PET)":                            "imaging",
    "Deductible":                                        "rx_ded",
    "Generic / Preferred / Non-preferred / Specialty":   "rx_tiers",
    "Single":                                            "rate_ee",
    "EE Only":                                           "rate_ee",
    "Employee Only":                                     "rate_ee",
    "Employee + Spouse":                                 "rate_es",
    "EE + Spouse":                                       "rate_es",
    "Employee + Child(ren)":                             "rate_ec",
    "Employee + Child":                                  "rate_ec",
    "EE + Child(ren)":                                   "rate_ec",
    "EE + Child":                                        "rate_ec",
    "Family":                                            "rate_fam",
    "Employee + Family":                                 "rate_fam",
    "EE + Family":                                       "rate_fam",
    "Rate Guarantee":                                    "notes",
}

def build_value(field_key, plan):
    if field_key == "ded":
        return f"{plan.get('ded_ind','')} / {plan.get('ded_fam','')}".strip(" /")
    if field_key == "moop":
        return f"{plan.get('moop_ind','')} / {plan.get('moop_fam','')}".strip(" /")
    if field_key == "rx_tiers":
        parts = [plan.get(k, "") for k in ("rx_generic", "rx_preferred", "rx_nonpref", "rx_specialty")]
        return " / ".join(p for p in parts if p)
    if field_key in ("rate_ee", "rate_es", "rate_ec", "rate_fam"):
        v = plan.get(field_key)
        try:
            return float(v) if v else ""
        except (TypeError, ValueError):
            return v or ""
    return plan.get(field_key, "") or ""

def find_template_anchors(ws):
    max_col = ws.max_column or 30
    start_col, carrier_row, plan_row = 8, -1, -1
    # Search broader range and lower starting col since templates may use earlier columns
    for r in range(8, 32):
        for c in range(4, min(max_col + 1, 35)):
            raw = ws.cell(r, c).value
            if raw is None: continue
            v = str(raw).strip().lower()
            if v in ("plan 1", "plan #1", "plan#1"):
                start_col, plan_row = c, r
                # Look up to 4 rows above for the carrier row
                for cr in range(r - 1, max(r - 5, 0), -1):
                    cv = ws.cell(cr, c).value
                    if cv is not None:
                        carrier_row = cr
                        break
                break
        if plan_row > 0: break
    # Fallback: if no Plan 1 found, look for "Carrier 1" or "Current"
    if plan_row < 0:
        for r in range(8, 32):
            for c in range(4, min(max_col + 1, 35)):
                raw = ws.cell(r, c).value
                if raw is None: continue
                v = str(raw).strip().lower()
                if v in ("carrier 1", "carrier1", "current"):
                    carrier_row = r
                    start_col = c
                    plan_row = r + 1
                    break
            if plan_row > 0: break
    if carrier_row > 0 and plan_row < 0:
        plan_row = carrier_row + 1
    if plan_row < 0:
        for r in range(10, 30):
            if ws.cell(r, 2).value == "h":
                plan_row, carrier_row = r, r - 1
                break
    return start_col, carrier_row, plan_row

def safe_set(ws, row, col, value):
    if row < 1 or col < 1: return
    for merged in list(ws.merged_cells.ranges):
        if (merged.min_row <= row <= merged.max_row and
                merged.min_col <= col <= merged.max_col and
                not (merged.min_row == row and merged.min_col == col)):
            ws.unmerge_cells(str(merged))
            break
    ws.cell(row, col).value = value

def copy_cell_style(src, dst):
    if src.has_style:
        dst.font = src.font.copy()
        dst.fill = src.fill.copy()
        dst.border = src.border.copy()
        dst.alignment = src.alignment.copy()
        dst.number_format = src.number_format

def write_excel_bundled(template_bytes, selected_bundles, client_name, renewal_bundle, quote_type, census=None):
    """
    selected_bundles: list of {carrier, plans: [plan_dict, plan_dict, ...]}
        where len(plans) == bundle_size for the quote_type (1/2/3)
    renewal_bundle: optional {carrier, plans: [...]} same structure
    quote_type: "single" / "double" / "triple"
    census: optional {ee, es, ec, fam} numbers for D34-D37
    """
    bundle_size = PLANS_PER_BUNDLE.get(quote_type, 1)

    wb = openpyxl.load_workbook(io.BytesIO(template_bytes))
    ws = next((wb[n] for n in wb.sheetnames if "medical" in n.lower()), wb.active)
    max_row = ws.max_row

    # Map row labels to row numbers — context-aware (In-Network vs Out-of-Network)
    # When we hit a section header, switch the prefix used for ambiguous fields like Deductible/MOOP
    row_map = {}
    section = "in"  # "in" or "oon"
    for r in range(1, max_row + 1):
        v = ws.cell(r, 3).value
        if not v:
            continue
        s = str(v).strip()
        s_low = s.lower()
        # Detect section transitions
        if "out-of-network" in s_low or "out of network" in s_low:
            section = "oon"
            continue
        if "in-network" in s_low or "in network" in s_low or "prescription" in s_low or "monthly rates" in s_low or "financial" in s_low:
            # Reset section on entering known In-Network/Rx/Rates blocks (Rx/Rates rows aren't ambiguous)
            if "out" not in s_low:
                section = "in" if "out" not in s_low else "oon"
            if "prescription" in s_low or "monthly rates" in s_low or "financial" in s_low:
                section = "in"  # Rx and rates use distinct field keys, doesn't matter
            else:
                section = "in"
            continue
        # Look up the label
        if s in LABEL_MAP:
            key = LABEL_MAP[s]
            # Only remap key based on section for the truly ambiguous fields
            if section == "oon":
                if key == "ded":         key = "oon_ded"
                elif key == "moop":      key = "oon_moop"
                elif key == "coinsurance": key = "oon_coins"
                elif key == "rx_ded":    key = "rx_ded"  # Rx rows usually appear before OON
            # Only set if not already mapped (first occurrence wins)
            if key not in row_map:
                row_map[key] = r

    start_col, carrier_row, plan_row = find_template_anchors(ws)
    if not selected_bundles and not renewal_bundle:
        raise ValueError("No plans selected for output.")

    # Prepend renewal as the leftmost carrier-bundle
    bundles = []
    if renewal_bundle:
        bundles.append({
            "carrier": renewal_bundle.get("carrier") or "Current Renewal",
            "plans": renewal_bundle.get("plans") or [],
            "_is_renewal": True,
        })
    bundles.extend(selected_bundles)

    # Write census (D34-D37) if provided
    if census:
        try:
            if "ee"  in census and census["ee"]  is not None: ws.cell(34, 4).value = float(census["ee"])
            if "es"  in census and census["es"]  is not None: ws.cell(35, 4).value = float(census["es"])
            if "ec"  in census and census["ec"]  is not None: ws.cell(36, 4).value = float(census["ec"])
            if "fam" in census and census["fam"] is not None: ws.cell(37, 4).value = float(census["fam"])
        except (TypeError, ValueError):
            pass

    tmpl_col = start_col

    # Each bundle takes `bundle_size` columns
    for b_idx, bundle in enumerate(bundles):
        bundle_first_col = start_col + b_idx * bundle_size
        bundle_last_col  = bundle_first_col + bundle_size - 1
        carrier_name = bundle.get("carrier") or "Unknown"
        plans = bundle.get("plans") or []

        # Pad plans list to bundle_size
        while len(plans) < bundle_size:
            plans.append({})

        # Write each plan into its sub-column
        for slot, plan in enumerate(plans[:bundle_size]):
            col = bundle_first_col + slot

            # Copy column width from template's first data column
            if col != tmpl_col:
                tmpl_dim = ws.column_dimensions[get_column_letter(tmpl_col)]
                ws.column_dimensions[get_column_letter(col)].width = tmpl_dim.width

            # Plan name row
            if plan_row > 0:
                if col != tmpl_col:
                    copy_cell_style(ws.cell(plan_row, tmpl_col), ws.cell(plan_row, col))
                safe_set(ws, plan_row, col, plan.get("plan_name") or "")

            # Benefit rows
            for field_key, row in row_map.items():
                value = build_value(field_key, plan)
                if value == "" or value is None:
                    continue
                if col != tmpl_col:
                    copy_cell_style(ws.cell(row, tmpl_col), ws.cell(row, col))
                safe_set(ws, row, col, value)

        # Carrier row: write name once, merge across bundle's columns
        if carrier_row > 0:
            # Unmerge any existing merges in this range
            for merged in list(ws.merged_cells.ranges):
                if (merged.min_row <= carrier_row <= merged.max_row and
                        merged.max_col >= bundle_first_col and merged.min_col <= bundle_last_col):
                    ws.unmerge_cells(str(merged))

            tmpl_cell = ws.cell(carrier_row, tmpl_col)
            for c in range(bundle_first_col, bundle_last_col + 1):
                copy_cell_style(tmpl_cell, ws.cell(carrier_row, c))

            ws.cell(carrier_row, bundle_first_col).value = carrier_name
            if bundle_last_col > bundle_first_col:
                ws.merge_cells(start_row=carrier_row, start_column=bundle_first_col,
                               end_row=carrier_row, end_column=bundle_last_col)

    # Clear ALL leftover content (including styled-empty placeholder columns) in unused slots
    last_used_col = start_col + len(bundles) * bundle_size - 1
    from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
    blank_fill = PatternFill(fill_type=None)
    blank_font = Font()
    blank_border = Border()
    blank_alignment = Alignment()
    for c in range(last_used_col + 1, ws.max_column + 1):
        # First, unmerge any merges that touch this column above row plan_row+1 (header area)
        # and below in benefit area
        for merged in list(ws.merged_cells.ranges):
            # Only unmerge ranges that contain this column AND are entirely within columns >= last_used_col+1
            # (don't unmerge label columns C, D)
            if merged.min_col >= last_used_col + 1 and merged.min_col <= c <= merged.max_col:
                ws.unmerge_cells(str(merged))

        # Clear ALL cells in this column from carrier_row down through the data rows
        if carrier_row > 0:
            data_end = max(row_map.values()) if row_map else carrier_row + 30
            for r in range(carrier_row, data_end + 1):
                cell = ws.cell(r, c)
                cell.value = None
                # Strip styling so it doesn't look like a populated column
                cell.fill = blank_fill
                cell.font = blank_font
                cell.border = blank_border
                cell.alignment = blank_alignment

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def write_excel(template_bytes, selected_plans, client_name, renewal_data=None):
    wb = openpyxl.load_workbook(io.BytesIO(template_bytes))
    ws = next((wb[n] for n in wb.sheetnames if "medical" in n.lower()), wb.active)
    max_row = ws.max_row

    row_map = {}
    for r in range(1, max_row + 1):
        v = ws.cell(r, 3).value
        if v and str(v).strip() in LABEL_MAP:
            row_map[LABEL_MAP[str(v).strip()]] = r

    start_col, carrier_row, plan_row = find_template_anchors(ws)
    if not selected_plans:
        raise ValueError("No plans selected for output.")

    if renewal_data:
        renewal_plan_dict = {k: v for k, v in renewal_data.items() if k != "carrier"}
        # Don't use "Current Plan" as a fallback - leave empty if no name extracted
        # (user can edit the preview before generating)
        selected_plans = [{
            "carrier": renewal_data.get("carrier") or "Current Renewal",
            "plan": renewal_plan_dict,
            "_is_renewal": True
        }] + list(selected_plans)

    tmpl_col = start_col
    # No special fill for renewal - blend in naturally with the template
    for idx, sp in enumerate(selected_plans):
        col = start_col + idx
        plan = sp["plan"]
        is_renewal = sp.get("_is_renewal", False)

        if idx > 0:
            tmpl_dim = ws.column_dimensions[get_column_letter(tmpl_col)]
            ws.column_dimensions[get_column_letter(col)].width = tmpl_dim.width

        if plan_row > 0:
            if idx > 0:
                copy_cell_style(ws.cell(plan_row, tmpl_col), ws.cell(plan_row, col))
            safe_set(ws, plan_row, col, plan.get("plan_name") or "")

        for field_key, row in row_map.items():
            value = build_value(field_key, plan)
            if value == "" or value is None:
                continue
            if idx > 0:
                copy_cell_style(ws.cell(row, tmpl_col), ws.cell(row, col))
            safe_set(ws, row, col, value)

    if carrier_row > 0:
        groups = []
        for idx, sp in enumerate(selected_plans):
            col = start_col + idx
            name = sp["carrier"]
            if not groups or groups[-1][0] != name:
                groups.append([name, col, col])
            else:
                groups[-1][2] = col

        for name, first_col, last_col in groups:
            for merged in list(ws.merged_cells.ranges):
                if (merged.min_row <= carrier_row <= merged.max_row and
                        merged.max_col >= first_col and merged.min_col <= last_col):
                    ws.unmerge_cells(str(merged))
            tmpl_cell = ws.cell(carrier_row, tmpl_col)
            for c in range(first_col, last_col + 1):
                copy_cell_style(tmpl_cell, ws.cell(carrier_row, c))
            ws.cell(carrier_row, first_col).value = name
            if last_col > first_col:
                ws.merge_cells(start_row=carrier_row, start_column=first_col,
                               end_row=carrier_row, end_column=last_col)

    # Clear any leftover placeholder columns beyond the data we wrote
    n_written = len(selected_plans)  # total columns written (including renewal)
    last_data_col = start_col + n_written - 1

    # Detect placeholder cells in carrier_row and plan_row beyond our data
    if plan_row > 0 and carrier_row > 0:
        for c in range(last_data_col + 1, ws.max_column + 1):
            # Check if this column has placeholder text we should clear
            for r in [carrier_row, plan_row]:
                cell_val = ws.cell(r, c).value
                if cell_val:
                    s = str(cell_val).lower().strip()
                    if any(k in s for k in ("carrier ", "plan ", "carrier#", "plan#")):
                        # Unmerge if needed, then clear
                        for merged in list(ws.merged_cells.ranges):
                            if (merged.min_row <= r <= merged.max_row and
                                    merged.min_col <= c <= merged.max_col):
                                ws.unmerge_cells(str(merged))
                                break
                        ws.cell(r, c).value = None

            # Also clear benefit row values in unused columns (in case template has demo data)
            for field_key, row in row_map.items():
                cell_val = ws.cell(row, c).value
                if cell_val is not None:
                    for merged in list(ws.merged_cells.ranges):
                        if (merged.min_row <= row <= merged.max_row and
                                merged.min_col <= c <= merged.max_col):
                            ws.unmerge_cells(str(merged))
                            break
                    ws.cell(row, c).value = None

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ─── ROUTES ───────────────────────────────────────────────────────────────────
@app.route("/")
def index():
    return send_from_directory("static", "index.html")

@app.route("/api/auth/check")
def auth_check():
    return jsonify({"authed": bool(session.get("authed"))})

@app.route("/api/auth/login", methods=["POST"])
def login():
    pw = (request.get_json() or {}).get("password", "")
    if pw == APP_PASSWORD:
        session["authed"] = True
        session.permanent = True
        return jsonify({"ok": True})
    return jsonify({"error": {"message": "Incorrect password"}}), 401

@app.route("/api/auth/logout", methods=["POST"])
def logout():
    session.clear()
    return jsonify({"ok": True})

@app.route("/api/templates", methods=["GET"])
@require_auth
def list_templates():
    """Return which template types are available."""
    out = []
    for qt, path in TEMPLATE_FILES.items():
        info = {"type": qt, "available": path.exists()}
        if path.exists():
            stat = path.stat()
            info["sizeBytes"] = stat.st_size
            info["modifiedAt"] = int(stat.st_mtime)
        out.append(info)
    return jsonify({"templates": out})

@app.route("/api/templates/<quote_type>", methods=["POST"])
@require_auth
def upload_template(quote_type):
    qt = quote_type.lower()
    if qt not in TEMPLATE_FILES:
        return jsonify({"error": {"message": f"Invalid quote type: {quote_type}"}}), 400

    if "file" not in request.files:
        return jsonify({"error": {"message": "No file provided"}}), 400

    f = request.files["file"]
    if not f.filename.lower().endswith((".xlsx", ".xls")):
        return jsonify({"error": {"message": "Must be an .xlsx file"}}), 400

    target = TEMPLATE_FILES[qt]
    target.parent.mkdir(parents=True, exist_ok=True)
    f.save(str(target))
    return jsonify({"ok": True, "type": qt})

@app.route("/api/templates/<quote_type>", methods=["DELETE"])
@require_auth
def delete_template(quote_type):
    qt = quote_type.lower()
    if qt not in TEMPLATE_FILES:
        return jsonify({"error": {"message": f"Invalid quote type: {quote_type}"}}), 400
    if TEMPLATE_FILES[qt].exists():
        TEMPLATE_FILES[qt].unlink()
    return jsonify({"ok": True})

@app.route("/api/templates/<quote_type>/download", methods=["GET"])
@require_auth
def download_template(quote_type):
    qt = quote_type.lower()
    if qt not in TEMPLATE_FILES or not TEMPLATE_FILES[qt].exists():
        return jsonify({"error": {"message": "Template not found"}}), 404
    return send_file(
        str(TEMPLATE_FILES[qt]),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"template_{qt}.xlsx"
    )

@app.route("/api/extract", methods=["POST"])
@require_auth
def extract():
    if not ANTHROPIC_KEY:
        return jsonify({"error": {"message": "ANTHROPIC_API_KEY not configured on server."}}), 500

    payload = request.get_json()
    page_range = payload.pop("_pageRange", "all")

    try:
        pdf_b64 = payload["messages"][0]["content"][0]["source"]["data"]
        if page_range and page_range.strip().lower() not in ("", "all"):
            pdf_bytes = base64.b64decode(pdf_b64)
            filtered = filter_pdf_pages(pdf_bytes, page_range)
            payload["messages"][0]["content"][0]["source"]["data"] = base64.b64encode(filtered).decode()
    except (KeyError, IndexError):
        pass

    try:
        req = urllib.request.Request(
            "https://api.anthropic.com/v1/messages",
            data=json.dumps(payload).encode(),
            headers={
                "Content-Type": "application/json",
                "x-api-key": ANTHROPIC_KEY,
                "anthropic-version": "2023-06-01",
            },
            method="POST"
        )
        with urllib.request.urlopen(req, timeout=120) as resp:
            return app.response_class(response=resp.read(), status=200, mimetype="application/json")
    except urllib.error.HTTPError as e:
        return app.response_class(response=e.read(), status=e.code, mimetype="application/json")
    except Exception as e:
        return jsonify({"error": {"message": str(e)}}), 500


@app.route("/api/generate", methods=["POST"])
@require_auth
def generate():
    payload = request.get_json()
    # selected_bundles: list of {carrier, plans: [plan_dict, ...]} where plans length == bundle_size
    selected_bundles = payload.get("selectedBundles", [])
    client_name = payload.get("clientName", "Group")
    renewal_bundle = payload.get("renewalBundle")  # {carrier, plans: [...]}
    quote_type = (payload.get("quoteType") or "single").lower()
    census = payload.get("census")  # {ee, es, ec, fam} optional

    template_path = pick_template_path(quote_type)
    available = [qt for qt, p in TEMPLATE_FILES.items() if p.exists()]
    print(f"[generate] quote_type={quote_type}, bundles={len(selected_bundles)}, renewal={bool(renewal_bundle)}, available_templates={available}")

    if not template_path or not template_path.exists():
        return jsonify({"error": {"message": f"No '{quote_type}' template uploaded. Upload one via Manage Templates."}}), 400

    try:
        template_bytes = template_path.read_bytes()
        buf = write_excel_bundled(template_bytes, selected_bundles, client_name, renewal_bundle, quote_type, census)
        safe_name = client_name.replace(" ", "_") or "Group"
        return send_file(
            buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f"{safe_name}_Medical_Comparison.xlsx"
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": {"message": str(e)}}), 500


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 8742))
    app.run(host="0.0.0.0", port=port)
